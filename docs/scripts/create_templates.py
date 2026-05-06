"""生成各表导入模板"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# 样式定义
HEADER_FILL = PatternFill("solid", fgColor="028090")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
REQUIRED_FONT = Font(bold=True, color="E74C3C", size=11)  # 必填字段红色
OPTIONAL_FONT = Font(color="666666", size=10)  # 可选字段灰色
THIN_BORDER = Border(
    left=Side(style="thin", color="D9DEE7"),
    right=Side(style="thin", color="D9DEE7"),
    top=Side(style="thin", color="D9DEE7"),
    bottom=Side(style="thin", color="D9DEE7"),
)

def create_template(path, headers, comments=None, sample_rows=None):
    """创建模板文件
    
    Args:
        path: 输出路径
        headers: 列定义 [(字段名, 是否必填, 字段说明), ...]
        comments: 列注释 {字段名: 注释}
        sample_rows: 示例数据行
    """
    wb = Workbook()
    ws = wb.active
    
    # 写表头
    for col, (name, required, desc) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
        # 设置列宽
        ws.column_dimensions[get_column_letter(col)].width = max(12, len(name) * 2)
    
    # 写字段说明行
    for col, (name, required, desc) in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=f"{'*必填' if required else '可选'}: {desc}")
        cell.font = REQUIRED_FONT if required else OPTIONAL_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
    
    # 写示例数据
    if sample_rows:
        for row_idx, row_data in enumerate(sample_rows, 3):
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.border = THIN_BORDER
    
    # 冻结首行
    ws.freeze_panes = "A2"
    
    wb.save(path)
    print(f"Created: {path}")

# 输出目录
OUT_DIR = "/sessions/69e8324fe6cf3c9b2576c7e7/workspace/demo-tauri/交付物/导入模板"

# 1. 会员表
create_template(
    f"{OUT_DIR}/01_member_会员表.xlsx",
    headers=[
        ("card_name", True, "会员卡名称"),
        ("phone", False, "电话号码"),
        ("name_remark", False, "名称备注"),
        ("coach", False, "教练"),
        ("remark", False, "备注"),
        ("prev_init_hours", False, "上月初始课时数(整数)"),
        ("month_new_hours", False, "本月新增课时数(整数)"),
        ("month_used_hours", False, "本月消耗课时数(整数)"),
        ("remain_hours", False, "剩余课时数(整数)"),
        ("churn_note", False, "流失说明"),
    ],
    sample_rows=[
        ["年卡-张三", "13800138001", "张三", "李教练", "续费用户", 20, 10, 5, 25, ""],
        ["季卡-李四", "13800138002", "李四", "王教练", "", 15, 0, 8, 7, ""],
    ]
)

# 2. 教练表
create_template(
    f"{OUT_DIR}/02_coach_教练表.xlsx",
    headers=[
        ("name", True, "教练姓名"),
        ("phone", False, "联系电话"),
        ("qualification", False, "资质证书"),
        ("store_name", False, "所属门店"),
        ("hourly_rate", False, "课时费单价(元)"),
        ("share_ratio", False, "分成比例(百分比,如50)"),
        ("status", False, "状态(1在职/0离职,默认1)"),
    ],
    sample_rows=[
        ["李教练", "13900139001", "国家一级教练", "总店", 200, 50, 1],
        ["王教练", "13900139002", "", "分店", 180, 45, 1],
    ]
)

# 3. 课程表
create_template(
    f"{OUT_DIR}/03_course_课程表.xlsx",
    headers=[
        ("coach_id", True, "教练ID(关联coach表)"),
        ("name", True, "课程名称"),
        ("description", False, "课程详情"),
        ("start_time", True, "上课时间(YYYY-MM-DD HH:MM:SS)"),
        ("end_time", False, "下课时间(YYYY-MM-DD HH:MM:SS)"),
        ("duration", False, "课时时长(分钟,默认60)"),
        ("max_students", False, "最大约课人数"),
        ("course_value", False, "单课价值(元)"),
        ("status", False, "状态(1进行中/2已结束/3已取消,默认1)"),
    ],
    sample_rows=[
        [1, "瑜伽基础班", "适合初学者", "2026-04-01 09:00:00", "2026-04-01 10:00:00", 60, 10, 100, 1],
        [2, "动感单车", "", "2026-04-01 14:00:00", "2026-04-01 15:00:00", 60, 15, 120, 1],
    ]
)

# 4. 约课记录表
create_template(
    f"{OUT_DIR}/04_booking_约课记录表.xlsx",
    headers=[
        ("course_id", True, "课程ID(关联course表)"),
        ("member_id", False, "会员ID(关联member表)"),
        ("member_card", False, "使用的会员卡"),
        ("member_name", True, "约课人姓名"),
        ("booking_time", False, "约课时间(YYYY-MM-DD HH:MM:SS)"),
        ("status", False, "状态(1已约/2已签到/3已取消,默认1)"),
    ],
    sample_rows=[
        [1, 1, "年卡-张三", "张三", "2026-04-01 08:30:00", 1],
        [1, 2, "季卡-李四", "李四", "2026-04-01 08:45:00", 2],
    ]
)

# 5. 课时费记录表
create_template(
    f"{OUT_DIR}/05_coach_fee_课时费记录表.xlsx",
    headers=[
        ("coach_id", True, "教练ID"),
        ("course_id", True, "课程ID"),
        ("actual_students", False, "实际约课人数"),
        ("course_value", False, "单课价值(元)"),
        ("share_ratio", False, "分成比例(百分比)"),
        ("total_fee", False, "总课时费(元)"),
        ("fee_formula", False, "计算规则说明"),
        ("settle_month", True, "结算月份(YYYY-MM)"),
        ("status", False, "状态(1待结算/2已结算,默认1)"),
    ],
    sample_rows=[
        [1, 1, 8, 100, 50, 400, "单课价值×人数×分成比例", "2026-04", 1],
    ]
)

# 9. 授课日记
create_template(
    f"{OUT_DIR}/09_teaching_diary_授课日记.xlsx",
    headers=[
        ("coach_id", False, "教练ID(可空)"),
        ("coach_name", True, "教练姓名"),
        ("course_name", True, "课程名称"),
        ("class_date", True, "上课日期(YYYY-MM-DD)"),
        ("start_time", False, "上课时间(HH:MM)"),
        ("duration", False, "时长(分钟,默认60)"),
        ("students", False, "学员数"),
        ("unit_value", False, "单课价值(元)"),
        ("remark", False, "备注"),
    ],
    sample_rows=[
        [1, "李教练", "瑜伽基础班", "2026-04-01", "09:00", 60, 8, 100, ""],
        [2, "王教练", "动感单车", "2026-04-01", "14:00", 60, 12, 120, ""],
    ]
)

# 10. 日常开支
create_template(
    f"{OUT_DIR}/10_daily_expense_日常开支.xlsx",
    headers=[
        ("expense_date", True, "支出日期(YYYY-MM-DD)"),
        ("expense_type", True, "支出类型"),
        ("project", True, "项目明细"),
        ("amount", True, "支出金额(元)"),
        ("remark", False, "备注"),
    ],
    sample_rows=[
        ["2026-04-01", "房租", "4月房租", 5000, ""],
        ["2026-04-05", "水电", "4月水电费", 800, ""],
        ["2026-04-10", "办公用品", "打印纸、文具", 200, ""],
    ]
)

# 11. 体验课
create_template(
    f"{OUT_DIR}/11_trial_class_体验课.xlsx",
    headers=[
        ("class_date", True, "体验日期(YYYY-MM-DD)"),
        ("trial_type", True, "体验课类型"),
        ("student_name", True, "体验课学员"),
        ("source", False, "来源"),
        ("unit_value", False, "体验课价值(元)"),
        ("coach_name", True, "教练"),
        ("status", False, "状态(1待体验/2已体验/3已转化/4已流失,默认1)"),
        ("remark", False, "备注"),
    ],
    sample_rows=[
        ["2026-04-01", "瑜伽体验", "张小明", "朋友推荐", 50, "李教练", 2, ""],
        ["2026-04-05", "动感单车体验", "王小红", "抖音", 60, "王教练", 3, "已办卡"],
    ]
)

# 12. 请假记录
create_template(
    f"{OUT_DIR}/12_coach_leave_请假记录.xlsx",
    headers=[
        ("coach_id", False, "教练ID(可空)"),
        ("coach_name", True, "教练姓名"),
        ("leave_month", True, "月份(YYYY-MM)"),
        ("leave_days", True, "请假天数(可填0.5)"),
        ("remark", False, "备注"),
    ],
    sample_rows=[
        [1, "李教练", "2026-04", 2, "事假"],
        [2, "王教练", "2026-04", 0.5, "半天病假"],
    ]
)

print("\n所有导入模板已生成完成！")
